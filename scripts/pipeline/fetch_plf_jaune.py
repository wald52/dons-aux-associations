"""Télécharge les CSV officiels de l'annexe Jaune « Effort financier de l'État
en faveur des associations », toutes années disponibles.

Pourquoi repartir de la source plutôt que des fichiers `data/sources/*.js`
déjà convertis : les convertisseurs historiques ont perdu des colonnes, et pas
les mêmes selon l'année. Les PLF 2018, 2019 et 2020 (218 667 lignes) ont perdu
à la fois le département et le RNA, que le CSV amont contient pourtant. Repartir
de l'amont les récupère.

Découverte : l'API data.gouv.fr est interrogée sur les deux intitulés
successifs de l'annexe (« associations subventionnées » jusqu'au PLF 2014,
« effort financier de l'État » ensuite). Aucune liste d'URL n'est codée en dur,
pour que les millésimes à venir soient récupérés tout seuls.

Usage :
    python3 scripts/pipeline/fetch_plf_jaune.py [--force] [--only 2022]

Idempotent : ne retélécharge pas un fichier déjà présent et de taille cohérente.
Sortie : data/raw/plf-jaune/*.csv (non versionné) + manifest.json (versionné).
"""

import argparse
import hashlib
import io
import json
import os
import re
import sys
import time

import requests

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
RAW_DIR = os.path.join(ROOT, "data", "raw", "plf-jaune")
MANIFEST = os.path.join(ROOT, "data", "sources-manifest", "plf-jaune.json")

API = "https://www.data.gouv.fr/api/1/datasets/"
QUERIES = [
    "jaune associations subventionnees",
    "effort financier etat faveur des associations",
]
# Un dataset n'est retenu que s'il parle bien de l'annexe Jaune « associations »
# (les deux intitulés successifs) ET porte une année de PLF identifiable.
JAUNE_OK = re.compile(r"jaune|effort\s+financier", re.I)
ASSOC_OK = re.compile(r"association", re.I)
# « PLF 2022 », « PLF24 » (millésime sur 2 chiffres), « pour 2019 ».
PLF_YEAR_4 = re.compile(r"\bPLF[ _-]?(20\d{2})\b|pour\s+(20\d{2})", re.I)
PLF_YEAR_2 = re.compile(r"\bPLF[ _-]?(\d{2})\b", re.I)

SESSION = requests.Session()
SESSION.headers.update({"Accept": "application/json", "User-Agent": "dons-aux-associations/pipeline"})


def plf_year_of(title, slug):
    for text in (title, slug):
        m = PLF_YEAR_4.search(text or "")
        if m:
            return int(m.group(1) or m.group(2))
    for text in (title, slug):
        m = PLF_YEAR_2.search(text or "")
        if m:
            return 2000 + int(m.group(1))
    m = re.search(r"(20\d{2})", slug or "")
    return int(m.group(1)) if m else None


def candidates(ds):
    """Ressources téléchargeables, de la plus fiable à la moins fiable.

    Le fichier d'origine (« attachments ») prime sur l'export généré par le
    portail : pour plusieurs millésimes cet export est vide (records_count = 0
    côté ODS). Le XLSX sert de repli — pour les PLF 2019 et 2021 c'est le seul
    format qui porte réellement les données.
    """
    XLSX = "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    out = []
    for r in ds.get("resources", []):
        fmt = (r.get("format") or "").lower()
        url = r.get("url") or ""
        title = r.get("title") or ""
        if "fiche" in title.lower():
            continue  # fiche explicative : pas des données
        if fmt == "csv":
            rank = 0 if "/attachments/" in url else 2
        elif fmt == XLSX:
            rank = 1 if "/attachments/" in url else 3
        else:
            continue
        if "/explore/dataset/" in url:
            rank += 10  # ancienne URL de téléchargement, souvent en 404
        out.append({"rank": rank, "format": "xlsx" if fmt == XLSX else "csv",
                    "url": url, "title": title})
    out.sort(key=lambda x: x["rank"])
    return out


def discover():
    """Retourne {plf_year: {...}} — un jeu de données par millésime de PLF."""
    found = {}
    for q in QUERIES:
        page = 1
        while page <= 3:
            r = SESSION.get(API, params={"q": q, "page_size": 50, "page": page}, timeout=60)
            r.raise_for_status()
            payload = r.json()
            for ds in payload.get("data", []):
                title, slug = ds.get("title") or "", ds.get("slug") or ""
                if not (JAUNE_OK.search(title) and ASSOC_OK.search(title)):
                    continue
                year = plf_year_of(title, slug)
                if not year:
                    continue

                cands = candidates(ds)
                if not cands:
                    continue
                prev = found.get(year)
                if prev and len(prev["candidates"]) >= len(cands) and prev["candidates"][0]["rank"] <= cands[0]["rank"]:
                    continue
                found[year] = {
                    "plf_year": year,
                    "dataset_slug": slug,
                    "dataset_title": title,
                    "dataset_page": f"https://www.data.gouv.fr/datasets/{slug}/",
                    "license": ds.get("license"),
                    "candidates": cands,
                }
            if not payload.get("next_page"):
                break
            page += 1
    return dict(sorted(found.items()))


MIN_BYTES = 10240  # en deçà, le portail a renvoyé une coquille vide


def looks_like_data(path, fmt):
    """Un export vide renvoie 200 avec quelques octets : il faut le détecter."""
    if not os.path.exists(path) or os.path.getsize(path) < MIN_BYTES:
        return False
    if fmt == "csv":
        with open(path, "rb") as f:
            head = f.read(8192)
        first = head.split(b"\n", 1)[0]
        return b";" in first or b"," in first
    return True  # xlsx : la conversion validera


def xlsx_to_csv(xlsx_path, csv_path):
    """Convertit sans rien interpréter : on change le contenant, pas le contenu."""
    import csv as _csv
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    n = 0
    with open(csv_path, "w", encoding="cp1252", errors="replace", newline="") as f:
        w = _csv.writer(f, delimiter=";")
        for row in ws.iter_rows(values_only=True):
            if row is None:
                continue
            cells = ["" if c is None else str(c) for c in row]
            if not any(c.strip() for c in cells):
                continue
            w.writerow(cells)
            n += 1
    wb.close()
    return n


def fetch_url(url, dest):
    for attempt in range(4):
        try:
            with SESSION.get(url, stream=True, timeout=240) as r:
                r.raise_for_status()
                tmp = dest + ".part"
                with open(tmp, "wb") as f:
                    for chunk in r.iter_content(1 << 20):
                        f.write(chunk)
                os.replace(tmp, dest)
            return None
        except Exception as e:
            if attempt == 3:
                return str(e)[:160]
            time.sleep(2 ** attempt)
    return "inatteignable"


def download(entry, force=False):
    """Essaie les candidats dans l'ordre jusqu'à obtenir un fichier exploitable."""
    os.makedirs(RAW_DIR, exist_ok=True)
    year = entry["plf_year"]
    path = os.path.join(RAW_DIR, f"plf-{year}.csv")

    if os.path.exists(path) and not force and looks_like_data(path, "csv"):
        entry["cached"] = True
    else:
        errors = []
        for cand in entry.get("candidates", []):
            raw_dest = path if cand["format"] == "csv" else os.path.join(RAW_DIR, f"plf-{year}.xlsx")
            err = fetch_url(cand["url"], raw_dest)
            if err:
                errors.append(f"{cand['format']}: {err}")
                continue
            if not looks_like_data(raw_dest, cand["format"]):
                errors.append(f"{cand['format']}: fichier vide ou non tabulaire")
                try:
                    os.remove(raw_dest)
                except OSError:
                    pass
                continue
            if cand["format"] == "xlsx":
                try:
                    rows = xlsx_to_csv(raw_dest, path)
                except Exception as e:
                    errors.append(f"xlsx: conversion impossible ({str(e)[:80]})")
                    continue
                if rows < 2:
                    errors.append("xlsx: aucune ligne")
                    continue
                entry["converted_from"] = "xlsx"
            entry["source_url"] = cand["url"]
            entry["source_format"] = cand["format"]
            entry["cached"] = False
            break
        else:
            entry["error"] = " | ".join(errors) or "aucun candidat"
            return entry

    if os.path.exists(path):
        h = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(1 << 20), b""):
                h.update(chunk)
        entry["file"] = os.path.relpath(path, ROOT)
        entry["bytes"] = os.path.getsize(path)
        entry["sha256"] = h.hexdigest()
        with open(path, "rb") as f:
            entry["lines"] = max(0, sum(1 for _ in f) - 1)
    return entry


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--force", action="store_true", help="retélécharger même si présent")
    ap.add_argument("--only", type=int, help="un seul millésime de PLF")
    args = ap.parse_args()

    print("Découverte des jeux de données PLF Jaune sur data.gouv.fr")
    found = discover()
    if args.only:
        found = {k: v for k, v in found.items() if k == args.only}
    print(f"  {len(found)} millésimes : {', '.join(str(y) for y in found)}\n")

    entries = []
    for year, entry in found.items():
        e = download(entry, force=args.force)
        entries.append(e)
        if e.get("error"):
            print(f"  PLF {year}  ÉCHEC  {e['error']}")
        else:
            tag = "cache" if e.get("cached") else "téléchargé"
            print(f"  PLF {year}  {e['bytes']/1048576:7.1f} Mo  {e.get('lines', 0):>7} lignes  ({tag})")

    os.makedirs(os.path.dirname(MANIFEST), exist_ok=True)
    with open(MANIFEST, "w", encoding="utf-8") as f:
        json.dump({
            "family": "plf_jaune",
            "fetched_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            "discovery_queries": QUERIES,
            "datasets": entries,
        }, f, ensure_ascii=False, indent=2)
        f.write("\n")

    ok = [e for e in entries if not e.get("error")]
    print(f"\n  {len(ok)}/{len(entries)} récupérés — "
          f"{sum(e.get('lines', 0) for e in ok)} lignes brutes")
    print(f"  -> {os.path.relpath(MANIFEST, ROOT)}")


if __name__ == "__main__":
    main()
