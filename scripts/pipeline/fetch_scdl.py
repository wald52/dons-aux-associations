"""Moissonneur générique des données de subventions publiées au standard SCDL.

C'est le levier d'exhaustivité du projet. Les 152 sources héritées avaient été
converties une par une, à la main ; ce script interroge l'API data.gouv.fr,
récupère tout ce qui ressemble à des subventions, et **vérifie les colonnes
réellement présentes** avant de retenir un fichier. Un millésime publié demain
par n'importe quelle collectivité sera pris sans modification du code.

Découverte, par union de plusieurs angles (aucun ne suffit seul) :
  - `schema=scdl/subventions` : les jeux qui déclarent le standard (~53) ;
  - les tags `subvention` / `subventions` (~330 et ~240) ;
  - quelques recherches plein texte.

Validation : un CSV n'est retenu que s'il porte au moins une colonne de
bénéficiaire ET une colonne de montant reconnues. C'est ce qui écarte les
subventions agricoles, les délibérations et les pages HTML déguisées en CSV
que les trois angles ci-dessus ramènent inévitablement.

Usage :
    python3 scripts/pipeline/fetch_scdl.py [--limite N] [--force] [--workers 8]

Idempotent : un fichier déjà téléchargé et valide n'est pas repris.
Sortie : data/raw/scdl/*.csv (non versionné) + data/sources-manifest/scdl.json
"""

import argparse
import concurrent.futures as cf
import hashlib
import io
import json
import os
import re
import sys
import time
import urllib.parse

import requests

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import common as C

ROOT = C.ROOT
RAW = os.path.join(ROOT, "data", "raw", "scdl")
MANIFEST = os.path.join(ROOT, "data", "sources-manifest", "scdl.json")
API = "https://www.data.gouv.fr/api/1/datasets/"

DECOUVERTE = [
    {"schema": "scdl/subventions"},
    {"tag": "subvention"},
    {"tag": "subventions"},
    {"q": "subventions associations"},
    {"q": "données essentielles subventions"},
    {"q": "subventions versées associations"},
]

# La reconnaissance des colonnes vit dans `common.py` : elle est partagée avec
# le moissonneur Opendatasoft et les normaliseurs, pour qu'un élargissement
# profite à tout le pipeline d'un coup.

MAX_OCTETS = 80 * 1024 * 1024     # au-delà, ce n'est pas une liste de subventions
MIN_OCTETS = 120                  # en deçà, coquille vide

SESSION = requests.Session()
SESSION.headers.update({"Accept": "application/json", "User-Agent": "dons-aux-associations/pipeline"})


def xlsx_vers_csv(source, destination):
    """Convertit la première feuille d'un classeur en CSV, sans interprétation."""
    import csv as _csv
    from openpyxl import load_workbook
    wb = load_workbook(source, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    n = 0
    with open(destination, "w", encoding="utf-8", newline="") as f:
        w = _csv.writer(f, delimiter=";")
        for ligne in ws.iter_rows(values_only=True):
            if ligne is None:
                continue
            cases = ["" if c is None else str(c) for c in ligne]
            if not any(c.strip() for c in cases):
                continue
            w.writerow(cases)
            n += 1
    wb.close()
    return n


def entete_valide(colonnes):
    """(vrai/faux, raison) — le fichier porte-t-il bien des subventions ?"""
    return C.porte_des_subventions(colonnes)


def decouvrir(limite=None):
    """Retourne {dataset_id: dataset} — union des angles de découverte."""
    vus = {}
    for params in DECOUVERTE:
        page, angle = 1, list(params)[0] + "=" + list(params.values())[0]
        while page <= 40:
            try:
                r = SESSION.get(API, params=dict(params, page_size=100, page=page), timeout=90)
                r.raise_for_status()
                charge = r.json()
            except Exception as e:
                print(f"    angle {angle} page {page} : {str(e)[:60]}")
                break
            for ds in charge.get("data", []):
                vus.setdefault(ds["id"], ds)
            if not charge.get("next_page"):
                break
            page += 1
        print(f"    {angle:34s} -> {len(vus):>5} jeux cumulés")
        if limite and len(vus) >= limite:
            break
    return vus


def ressources_csv(ds):
    """Ressources plausibles, une par fichier réel, avec toutes ses adresses.

    Un portail moissonné par data.gouv.fr ré-inscrit ses fichiers à chaque
    passage : le jeu de Grenoble-Alpes Métropole porte **1 044 ressources pour
    neuf fichiers**, la même année revenant sous des dizaines d'identifiants et
    d'adresses. Les prendre une par une, c'est télécharger cent fois le même
    CSV, puis le faire entrer cent fois dans la table sous cent `source_id`
    différents — 91 sources fantômes pour la seule métropole, dont 85 ne
    portaient plus qu'une ligne après déduplication.

    On regroupe donc les ressources par nom de fichier, et on garde leurs
    adresses de la plus récente à la plus ancienne. Aucune n'est privilégiée :
    la plus récente n'est pas forcément la bonne (le portail réorganise ses
    chemins sans que data.gouv.fr le sache), c'est `traiter_dataset` qui
    essaiera les suivantes tant qu'une adresse répond 404.
    """
    groupes = {}
    ordonnees = sorted(ds.get("resources", []),
                       key=lambda r: r.get("last_modified") or r.get("created_at") or "",
                       reverse=True)
    for r in ordonnees:
        fmt = (r.get("format") or "").lower()
        url = r.get("url") or ""
        if fmt not in ("csv", "xlsx", "ods"):
            continue
        # Une « ressource » pointant vers le site de la collectivité est parfois
        # une page d'information, pas un jeu de données — d'où ce filtre. Mais
        # exiger que l'adresse FINISSE par « .csv » écartait tous les points
        # d'export d'API, qui servent pourtant de vrais fichiers :
        #
        #   .../api/explore/v2.1/catalog/datasets/<jeu>/exports/csv
        #   .../dataset/datasets/388/resource/493/download/
        #
        # Mesuré le 21/08/2026 sur les six angles de découverte : **333 jeux de
        # 63 organisations** étaient perdus ainsi, SANS AUCUNE TRACE dans le
        # manifeste — ni retenus, ni écartés, invisibles. Parmi eux, des
        # collectivités que le site ne couvre pas du tout : les départements de
        # l'Aude et des Hauts-de-Seine, Grand Paris Seine Ouest, Bourges Plus,
        # la Région Pays de la Loire.
        #
        # On fait donc confiance au FORMAT déclaré quand l'adresse ressemble à
        # un téléchargement. Ce qui répond du HTML est écarté plus loin, avec
        # son vrai motif — un rejet doit dire pourquoi.
        hote = urllib.parse.urlparse(url).netloc
        chemin = url.lower().split("?")[0]
        telechargeable = (chemin.endswith((".csv", ".xlsx", ".ods"))
                          or chemin.rstrip("/").endswith(("/exports/csv", "/exports/xlsx",
                                                          "/exports/ods", "/download"))
                          or "/download/" in chemin
                          or "/exports/" in chemin)
        if hote not in ("static.data.gouv.fr", "www.data.gouv.fr", "data.gouv.fr") \
           and not telechargeable:
            continue
        # On ne regroupe QUE les fichiers hébergés par la collectivité :
        # c'est là que vit la ré-inscription. Sur `static.data.gouv.fr`, chaque
        # ressource est un dépôt distinct — deux millésimes y portent souvent
        # le même nom de fichier, et les fondre effacerait une année entière.
        base = urllib.parse.unquote(url.split("?")[0].rsplit("/", 1)[-1]).lower()
        if hote.endswith("data.gouv.fr") or not base.endswith((".csv", ".xlsx", ".ods")):
            cle = ("id", r.get("id") or url)
        else:
            cle = (fmt, base, C.fold(r.get("title") or ""))
        if cle not in groupes:
            groupes[cle] = {"id": r.get("id"), "titre": r.get("title") or "",
                            "format": fmt, "url": url, "urls": []}
        if url not in groupes[cle]["urls"]:
            groupes[cle]["urls"].append(url)
    return list(groupes.values())


def telecharger(urls, destination):
    """Écrit la première adresse qui répond ; retourne celle qui a servi.

    Un portail qui réorganise ses chemins laisse derrière lui des inscriptions
    périmées chez data.gouv.fr : les neuf fichiers de Grenoble étaient déclarés
    sous des adresses datées qui répondent toutes 404, alors que la forme
    courante est servie normalement. Essayer les adresses connues du fichier,
    de la plus récente à la plus ancienne, récupère le fichier sans avoir à
    deviner la règle de réécriture du portail.
    """
    derniere = None
    for url in urls:
        try:
            with SESSION.get(url, stream=True, timeout=120) as rep:
                rep.raise_for_status()
                if int(rep.headers.get("content-length") or 0) > MAX_OCTETS:
                    raise ValueError("fichier trop volumineux")
                tmp = destination + ".part"
                recu = 0
                with open(tmp, "wb") as f:
                    for bloc in rep.iter_content(1 << 18):
                        recu += len(bloc)
                        if recu > MAX_OCTETS:
                            break
                        f.write(bloc)
                os.replace(tmp, destination)
            return url
        except Exception as e:
            derniere = e
    raise derniere or ValueError("aucune adresse")


def traiter_dataset(ds, force=False):
    """Télécharge et valide les ressources d'un jeu. Retourne une fiche."""
    org = ds.get("organization") or {}
    fiche = {
        "dataset_id": ds["id"], "slug": ds.get("slug"), "titre": ds.get("title"),
        "page": f"https://www.data.gouv.fr/datasets/{ds.get('slug')}/",
        "organisation": org.get("name"), "organisation_id": org.get("id"),
        "licence": ds.get("license"), "derniere_maj": ds.get("last_modified"),
        "fichiers": [], "ecartes": [],
    }
    # Dernier filet contre le même fichier publié deux fois sous deux noms :
    # le regroupement par nom de fichier ne voit pas un `subventions_2019.csv`
    # aussi déposé en `subventions-2019-v2.csv`. L'empreinte, elle, le voit.
    vus = {}
    for res in ressources_csv(ds):
        if res["format"] == "ods":
            fiche["ecartes"].append({"titre": res["titre"], "raison": "format ods"})
            continue
        nom = f"{ds['id'][:8]}-{(res['id'] or '')[:8]}.csv"
        chemin = os.path.join(RAW, nom)
        # Un tableur est converti en CSV sans rien interpréter : on change le
        # contenant, pas le contenu. Sans cela, 110 fichiers seraient perdus
        # pour la seule raison qu'ils sont publiés en XLSX.
        if res["format"] == "xlsx" and not (os.path.exists(chemin) and not force):
            brut = os.path.join(RAW, nom[:-4] + ".xlsx")
            try:
                res["url"] = telecharger(res["urls"], brut)
                if xlsx_vers_csv(brut, chemin) < 2:
                    raise ValueError("aucune ligne exploitable")
            except Exception as e:
                fiche["ecartes"].append({"titre": res["titre"], "raison": "xlsx : " + str(e)[:60]})
                for f_ in (brut, chemin):
                    if os.path.exists(f_):
                        os.remove(f_)
                continue
            finally:
                if os.path.exists(brut):
                    os.remove(brut)
        if res["format"] == "csv" and not (
                os.path.exists(chemin) and not force and os.path.getsize(chemin) > MIN_OCTETS):
            try:
                res["url"] = telecharger(res["urls"], chemin)
            except Exception as e:
                fiche["ecartes"].append({"titre": res["titre"], "raison": str(e)[:80]})
                continue

        if os.path.getsize(chemin) < MIN_OCTETS:
            fiche["ecartes"].append({"titre": res["titre"], "raison": "fichier vide"})
            os.remove(chemin)
            continue
        # Un point d'export qui répond une page d'erreur ou un écran de
        # connexion sert du HTML sous un format déclaré « csv ». Le dire ainsi
        # plutôt que de laisser `read_rows` échouer sur un motif obscur :
        # un rejet doit nommer sa vraie cause.
        with open(chemin, "rb") as f:
            debut = f.read(400).lstrip().lower()
        if debut.startswith((b"<!doctype", b"<html", b"<?xml")):
            fiche["ecartes"].append({"titre": res["titre"], "url": res["url"],
                                     "raison": "réponse HTML, pas un fichier de données"})
            os.remove(chemin)
            continue
        try:
            entete, _, meta = C.read_rows(chemin)
        except Exception as e:
            fiche["ecartes"].append({"titre": res["titre"], "raison": "illisible : " + str(e)[:60]})
            continue
        ok, raison = entete_valide(entete)
        if not ok:
            # On consigne les colonnes réelles : c'est ce qui permet de voir si
            # le rejet vient d'un fichier hors sujet ou d'une graphie que le
            # dictionnaire ne connaît pas encore.
            fiche["ecartes"].append({"titre": res["titre"], "raison": raison,
                                     "colonnes": entete[:16], "url": res["url"]})
            os.remove(chemin)
            continue

        with open(chemin, "rb") as f:
            sha = hashlib.sha256(f.read()).hexdigest()
        if sha in vus:
            fiche["ecartes"].append({"titre": res["titre"],
                                     "raison": "même contenu que " + vus[sha]})
            os.remove(chemin)
            continue
        vus[sha] = res["titre"] or os.path.basename(chemin)
        fiche["fichiers"].append({
            "resource_id": res["id"], "titre": res["titre"], "url": res["url"],
            "adresses_connues": len(res["urls"]),
            "fichier": os.path.relpath(chemin, ROOT), "octets": os.path.getsize(chemin),
            "sha256": sha, "encodage": meta["encoding"], "separateur": meta["delimiter"],
            "colonnes": entete,
        })
    return fiche


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limite", type=int, help="borner le nombre de jeux (mise au point)")
    ap.add_argument("--force", action="store_true")
    ap.add_argument("--workers", type=int, default=8)
    args = ap.parse_args()

    os.makedirs(RAW, exist_ok=True)
    print("Moissonnage SCDL — découverte sur data.gouv.fr\n")
    jeux = decouvrir(args.limite)
    liste = list(jeux.values())[: args.limite] if args.limite else list(jeux.values())
    print(f"\n  {len(liste)} jeux à examiner\n")

    fiches, faits = [], 0
    with cf.ThreadPoolExecutor(max_workers=args.workers) as ex:
        futurs = {ex.submit(traiter_dataset, ds, args.force): ds for ds in liste}
        for fut in cf.as_completed(futurs):
            faits += 1
            try:
                fiches.append(fut.result())
            except Exception as e:
                print(f"    échec : {str(e)[:70]}")
            if faits % 50 == 0:
                retenus = sum(len(f["fichiers"]) for f in fiches)
                print(f"    {faits}/{len(liste)} examinés, {retenus} fichiers retenus")

    retenues = [f for f in fiches if f["fichiers"]]
    nb_fichiers = sum(len(f["fichiers"]) for f in retenues)
    octets = sum(x["octets"] for f in retenues for x in f["fichiers"])

    os.makedirs(os.path.dirname(MANIFEST), exist_ok=True)
    with open(MANIFEST, "w", encoding="utf-8") as f:
        json.dump({
            "family": "scdl",
            "fetched_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            "angles_de_decouverte": DECOUVERTE,
            "jeux_examines": len(liste),
            "jeux_retenus": len(retenues),
            "fichiers_retenus": nb_fichiers,
            "datasets": sorted(retenues, key=lambda x: x.get("organisation") or ""),
            # Tous les écartés, y compris ceux de jeux par ailleurs retenus :
            # c'est le journal qui permet d'élargir le dictionnaire de colonnes.
            "ecartes": [{"organisation": f.get("organisation"), "page": f.get("page"),
                         "ecartes": f["ecartes"]}
                        for f in fiches if f["ecartes"]],
        }, f, ensure_ascii=False, indent=2)
        f.write("\n")

    raisons = {}
    for f in fiches:
        for e in f["ecartes"]:
            raisons[e["raison"][:44]] = raisons.get(e["raison"][:44], 0) + 1
    print(f"\n  jeux examinés .... {len(liste)}")
    print(f"  jeux retenus ..... {len(retenues)}")
    print(f"  fichiers retenus . {nb_fichiers} ({octets/1048576:.0f} Mo)")
    print("\n  principales raisons d'écarter :")
    for r, n in sorted(raisons.items(), key=lambda x: -x[1])[:8]:
        print(f"    {n:>5}  {r}")
    print(f"\n  -> {os.path.relpath(MANIFEST, ROOT)}")


if __name__ == "__main__":
    main()
