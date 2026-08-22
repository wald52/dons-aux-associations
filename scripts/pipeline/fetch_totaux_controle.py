"""Les totaux de contrôle : combien va-t-il aux associations, en tout ?

Le site additionne ce qu'il a trouvé. Il ne sait pas dire ce que cela
représente — 149,68 Md€ cumulés sur 2001-2027, est-ce beaucoup ? tout ? un
dixième ? Sans ordre de grandeur extérieur, le total affiché ne veut rien dire.

Les comptes nationaux de l'INSEE donnent cet ordre de grandeur, et lui seul est
officiel : l'opération **D751, « transferts courants aux ISBLSM »**, mesure ce
que les administrations publiques versent chaque année aux institutions sans but
lucratif au service des ménages. Série annuelle depuis 1949, base 2020.

TROIS RÉSERVES, sans lesquelles la comparaison ment :

  1. Le secteur S15 (ISBLSM) N'EST PAS « les associations » : il exclut celles
     qui sont reclassées en administrations publiques (S13) ou en sociétés
     (S11). Le périmètre est plus étroit que le sens courant du mot.
  2. Le total du site est un CUMUL pluriannuel ; D751 est un FLUX annuel. La
     comparaison n'a de sens qu'exercice par exercice, jamais en bloc.
  3. D751 ne comprend que les transferts COURANTS : les subventions
     d'investissement (D92) sont ailleurs, et la commande publique n'y est pas
     du tout.

Usage :
    python3 scripts/pipeline/fetch_totaux_controle.py [--force]

Sortie : data/raw/insee/*.xlsx (non versionné)
       + data/canonical/totaux-controle.json (versionné)
"""

import argparse
import io
import json
import os
import sys
import time

import requests

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl

import common as C

ROOT = C.ROOT
RAW = os.path.join(ROOT, "data", "raw", "insee")
OUT = os.path.join(ROOT, "data", "canonical", "totaux-controle.json")

# Deux tableaux, deux points de vue sur le même flux : ce que les
# administrations VERSENT (7.301, compte des APU) et ce que les associations
# REÇOIVENT (7.501, compte des ISBLSM). L'écart entre les deux, ce sont les
# ménages et les entreprises.
TABLEAUX = [
    {
        "id": "T_7301",
        "url": "https://www.insee.fr/fr/statistiques/fichier/8068612/T_7301.xlsx",
        "titre": "7.301 — Compte des administrations publiques (S13)",
        "section": "Emplois",
        "codes": {"D751": "Transferts courants aux ISBLSM versés par les APU"},
    },
    {
        "id": "T_7501",
        "url": "https://www.insee.fr/fr/statistiques/fichier/8068616/T_7501.xlsx",
        "titre": "7.501 — Compte des institutions sans but lucratif au service "
                 "des ménages (S15)",
        "section": "Ressources",
        "codes": {
            "D751": "Transferts courants aux ISBLSM reçus, toutes origines",
            # Le compte de capital suffixe ses codes : D9R « à recevoir »,
            # D9P « à payer ». Chercher « D92 » n'y trouve rien.
            "D92R": "Aides à l'investissement à recevoir",
            "D9R": "Transferts en capital à recevoir",
        },
    },
]

RESERVES = [
    "Le secteur S15 « institutions sans but lucratif au service des ménages » "
    "n'est pas l'ensemble des associations : celles qui sont reclassées en "
    "administrations publiques ou en sociétés en sont exclues.",
    "D751 est un flux ANNUEL ; le total affiché par le site est un cumul "
    "pluriannuel. La comparaison ne vaut qu'exercice par exercice.",
    "D751 ne couvre que les transferts courants : ni les subventions "
    "d'investissement, ni la commande publique.",
    "Les montants des comptes nationaux sont en milliards d'euros, arrondis au "
    "millième par la source.",
]

SESSION = requests.Session()
SESSION.headers["User-Agent"] = "dons-aux-associations/1.0 (+https://github.com/wald52)"


def telecharger(url, chemin, force=False):
    if os.path.exists(chemin) and not force:
        return True
    for essai in range(3):
        try:
            r = SESSION.get(url, timeout=180)
            r.raise_for_status()
            with open(chemin, "wb") as f:
                f.write(r.content)
            return True
        except Exception as e:
            print(f"      essai {essai+1} : {str(e)[:70]}")
            time.sleep(3 * (essai + 1))
    return False


def lire_tableau(chemin, section_voulue, codes):
    """Séries annuelles des codes demandés, dans la section voulue.

    Un même code apparaît DEUX FOIS dans ces tableaux — une fois en ressources,
    une fois en emplois. Prendre la première occurrence venue, c'est lire ce que
    les associations reçoivent là où on croyait lire ce que l'État verse. On
    suit donc l'intitulé de section (« Ressources » / « Emplois ») en descendant
    les lignes, et on ne retient que celles de la bonne section.
    """
    wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[-1]]
    annees = []
    section = None
    series = {}
    for ligne in ws.iter_rows(values_only=True):
        code = (str(ligne[0]).strip() if ligne[0] is not None else "")
        libelle = (str(ligne[1]).strip() if len(ligne) > 1 and ligne[1] is not None else "")
        if not annees and not code and libelle == "" and any(
                isinstance(c, (int, float)) or (isinstance(c, str) and c.strip().isdigit())
                for c in ligne[2:6]):
            annees = [str(c).strip() for c in ligne[2:] if c is not None]
            continue
        if not code and libelle in ("Ressources", "Emplois"):
            section = libelle
            continue
        if code in codes and section == section_voulue:
            valeurs = {}
            for i, v in enumerate(ligne[2:]):
                if i >= len(annees) or v is None or v == "":
                    continue
                try:
                    valeurs[annees[i]] = float(str(v).replace(",", "."))
                except ValueError:
                    continue
            series[code] = valeurs
    return series


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--force", action="store_true")
    args = ap.parse_args()
    os.makedirs(RAW, exist_ok=True)

    print("Totaux de contrôle — comptes nationaux de l'INSEE\n")
    sortie = {
        "generated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "unite": "milliards d'euros",
        "base": "Comptes nationaux annuels, base 2020",
        "reserves": RESERVES,
        "series": {},
    }
    for tab in TABLEAUX:
        chemin = os.path.join(RAW, tab["id"] + ".xlsx")
        print(f"  {tab['id']} — {tab['titre']}")
        if not telecharger(tab["url"], chemin, args.force):
            print("      ÉCHEC de téléchargement")
            continue
        series = lire_tableau(chemin, tab["section"], tab["codes"])
        for code, valeurs in series.items():
            if not valeurs:
                continue
            annees = sorted(valeurs)
            sortie["series"][f"{tab['id']}.{code}"] = {
                "tableau": tab["titre"],
                "url": tab["url"],
                "section": tab["section"],
                "code": code,
                "libelle": tab["codes"][code],
                "premiere_annee": annees[0],
                "derniere_annee": annees[-1],
                "valeurs_md_eur": {a: valeurs[a] for a in annees},
            }
            print(f"      {code:5s} {tab['codes'][code][:52]:54} "
                  f"{annees[0]}→{annees[-1]}  "
                  f"{valeurs[annees[-1]]:.3f} Md€ en {annees[-1]}")
        manquants = [c for c in tab["codes"] if c not in series]
        if manquants:
            print(f"      absents de la section {tab['section']} : {manquants}")

    if not sortie["series"]:
        print("\n  aucune série lue — rien n'est écrit")
        return 1
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(sortie, f, ensure_ascii=False, indent=1)
        f.write("\n")
    print(f"\n  -> {os.path.relpath(OUT, ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
